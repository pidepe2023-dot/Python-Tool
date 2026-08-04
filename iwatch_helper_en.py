"""
iWatch Filling Assistant
=============================
Reads an Excel file with several tabs (one per "master agent"). Inside
each tab it automatically detects the different contact blocks (for
example OWNER, COMPLIANCE OFFICER) and lets you choose which one to use.
Then it shows you each field with a "Copy" button so you can paste it
(Ctrl+V) into the iWatch form, or use auto-fill mode to fill everything.

INSTALLATION (on the work computer):
    pip install openpyxl pyperclip pyautogui

USAGE:
    python iwatch_helper_en.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import pyperclip
import threading
import time
import pyautogui

# pyautogui failsafe: if you move the mouse to the top-left corner of the
# screen at any moment, the auto-fill is ABORTED immediately (emergency stop).
pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.15  # small pause between each simulated action

# ---------------------------------------------------------------------------
# CONFIGURATION: adjust this list to the real field order / Tab in iWatch.
# "label" is how you want it shown on screen; "keywords" are the words the
# script will look for in the Excel label column (not case-sensitive) to
# find the matching value.
# ---------------------------------------------------------------------------
FIELD_MAP = [
    {"label": "First",              "keywords": ["FIRST"]},
    {"label": "Middle",              "keywords": ["MIDDLE"]},
    {"label": "Last",                "keywords": ["LAST"]},
    {"label": "Suffix",              "keywords": ["SUFFIX"]},
    {"label": "Agency",              "keywords": ["AGENCY"]},
    {"label": "Address 1",           "keywords": ["ADDRESS 1", "ADDRESS1"]},
    {"label": "Address 2",           "keywords": ["ADDRESS 2", "ADDRESS2"]},
    {"label": "Country",             "keywords": ["COUNTRY"]},
    {"label": "State",               "keywords": ["STATE"]},
    {"label": "City",                "keywords": ["CITY"]},
    {"label": "Zip",                 "keywords": ["ZIP"]},
    {"label": "Source",              "keywords": ["SOURCE"]},
    {"label": "International Code",  "keywords": ["INTERNATIONAL CODE"]},
    {"label": "Phone Number",        "keywords": ["PHONE NUMBER"]},
    {"label": "Extension",           "keywords": ["EXTENSION"]},
    {"label": "Additional Phone",    "keywords": ["ADDITIONAL PHONE"]},
    {"label": "Fax Number",          "keywords": ["FAX"]},
    {"label": "Email Address",       "keywords": ["EMAIL"]},
]

# Fields that are DROPDOWN MENUS (combobox) in iWatch instead of free text.
# Auto-fill pauses on these so you select them by hand, because pasting
# text into a combobox does not always work the same way as in a normal
# text field.
DROPDOWN_FIELDS = {"Country", "State"}


def _extract_fields_from_rows(ws, row_start, row_end):
    """Same as before, but limited to a row range (one block)."""
    found = {}
    for row in ws.iter_rows(min_row=row_start, max_row=row_end):
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value).strip().upper()
            if not text:
                continue
            for field in FIELD_MAP:
                if field["label"] in found:
                    continue
                for kw in field["keywords"]:
                    if kw == text or text.startswith(kw):
                        neighbor = ws.cell(row=cell.row, column=cell.column + 1)
                        value = neighbor.value
                        if value is not None and str(value).strip() != "":
                            found[field["label"]] = str(value).strip()
                        break
    return [(f["label"], found.get(f["label"], "")) for f in FIELD_MAP]


# All known keywords (used to know what is NOT a block title)
_ALL_KEYWORDS = {kw for f in FIELD_MAP for kw in f["keywords"]}


def _extract_fields_from_col(ws, row_start, row_end, label_col, value_col):
    """
    Like _extract_fields_from_rows, but ONLY looks at one label column
    (label_col). The value is normally in value_col (the column immediately
    to the right), but because of merged/offset cells (e.g. "INTERNATIONAL
    CODE" whose value lands 2 columns further instead of right next to it)
    it checks up to 3 columns to the right just in case.
    This is needed when there are two tables side by side in the same rows
    (e.g. Owner in columns B:C and Compliance Officer in columns F:G).
    """
    found = {}
    for r in range(row_start, row_end + 1):
        cell = ws.cell(row=r, column=label_col)
        if cell.value is None:
            continue
        text = str(cell.value).strip().upper()
        if not text:
            continue
        for field in FIELD_MAP:
            if field["label"] in found:
                continue
            for kw in field["keywords"]:
                if kw == text or text.startswith(kw):
                    # Look for the first non-empty value between value_col and value_col+2
                    for c in range(value_col, value_col + 3):
                        value = ws.cell(row=r, column=c).value
                        if value is not None and str(value).strip() != "":
                            found[field["label"]] = str(value).strip()
                            break
                    break
    return [(f["label"], found.get(f["label"], "")) for f in FIELD_MAP]


def extract_blocks_from_sheet(ws):
    """
    A single tab can have several contact blocks (OWNER, COMPLIANCE
    OFFICER, etc.), stacked vertically OR side by side in the same rows
    but different columns. This function:
      1. Finds EVERY cell where the "FIRST" label appears (row AND
         column) -> that starts a new block, in that column.
      2. For each one, figures out how far its table goes by scanning
         down that SAME column until finding 2 empty rows in a row (or
         the next "FIRST" in that same column, or a 30-row limit).
      3. Looks upward, near that column, for the block title (uppercase
         text that is not a known field label).
    Returns a list of dicts: {"name": "Owner", "fields": [...]}.
    If no "FIRST" is found, returns one generic block with the whole tab
    (so tabs with a different format do not break).
    """
    anchors = []  # list of (row, column)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value).strip().upper()
            if text == "FIRST":
                anchors.append((cell.row, cell.column))

    print(f"[DEBUG] Tab '{ws.title}': found {len(anchors)} 'FIRST' label(s) at: {anchors}")

    if not anchors:
        return [{"name": "General", "fields": _extract_fields_from_rows(ws, 1, ws.max_row)}]

    blocks = []
    for i, (start_row, label_col) in enumerate(anchors):
        value_col = label_col + 1

        # Limit: do not cross into the next "FIRST" that is in the SAME column
        same_col_next = [r for (r, c) in anchors if c == label_col and r > start_row]
        hard_limit = min(same_col_next) - 1 if same_col_next else ws.max_row
        hard_limit = min(hard_limit, start_row + 30, ws.max_row)

        # Scan down the column until 2 empty rows in a row (end of table)
        end_row = start_row
        empty_streak = 0
        r = start_row
        while r <= hard_limit:
            c1 = ws.cell(row=r, column=label_col).value
            c2 = ws.cell(row=r, column=value_col).value
            if (c1 is None or str(c1).strip() == "") and (c2 is None or str(c2).strip() == ""):
                empty_streak += 1
                if empty_streak >= 2:
                    break
            else:
                empty_streak = 0
                end_row = r
            r += 1

        # Look for the block title: scan up to 8 rows, checking nearby
        # columns (in case the title is centered/merged slightly differently)
        block_name = f"Block {i + 1}"
        for rr in range(start_row - 1, max(start_row - 9, 0), -1):
            found_name = None
            for cc in range(max(label_col - 1, 1), label_col + 3):
                cell = ws.cell(row=rr, column=cc)
                if cell.value is None:
                    continue
                text = str(cell.value).strip()
                if not text or text.upper() in _ALL_KEYWORDS:
                    continue
                if text.upper() == text and len(text) > 2:  # all uppercase
                    found_name = text.title()
                    break
            if found_name:
                block_name = found_name
                break

        blocks.append({
            "name": block_name,
            "fields": _extract_fields_from_col(ws, start_row, end_row, label_col, value_col),
        })

    print(f"[DEBUG] Detected blocks: {[b['name'] for b in blocks]}")
    return blocks


class IWatchHelper(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("iWatch Filling Assistant")
        self.geometry("520x600")
        self.minsize(480, 560)
        self.resizable(True, True)

        self.wb = None
        self.blocks = []  # list of {"name":.., "fields": [...]}
        self.fields = []  # list of (label, value) for the current block
        self.current_index = 0

        self._build_top_bar()
        self._build_field_view()

    # ---------------- UI ----------------
    def _build_top_bar(self):
        frame = ttk.Frame(self, padding=10)
        frame.pack(fill="x")

        ttk.Button(frame, text="Open Excel...", command=self.load_excel).pack(side="left")

        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(frame, textvariable=self.sheet_var, state="readonly", width=22)
        self.sheet_combo.pack(side="left", padx=8)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_selected)

        ttk.Label(frame, text="Role:").pack(side="left", padx=(6, 2))
        self.role_var = tk.StringVar()
        self.role_combo = ttk.Combobox(frame, textvariable=self.role_var, state="readonly", width=18)
        self.role_combo.pack(side="left")
        self.role_combo.bind("<<ComboboxSelected>>", self.on_role_selected)

    def _build_field_view(self):
        frame = ttk.Frame(self, padding=20)
        frame.pack(fill="both", expand=True)

        self.progress_label = ttk.Label(frame, text="", font=("Segoe UI", 9))
        self.progress_label.pack(anchor="w")

        self.field_label = ttk.Label(frame, text="Select a file and an agent to get started",
                                      font=("Segoe UI", 14, "bold"))
        self.field_label.pack(anchor="w", pady=(10, 4))

        self.value_box = tk.Text(frame, height=3, font=("Segoe UI", 12), wrap="word")
        self.value_box.pack(fill="x", pady=(0, 12))

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill="x")

        self.copy_btn = ttk.Button(btn_frame, text="Copy  (and advance)", command=self.copy_and_next)
        self.copy_btn.pack(side="left")

        ttk.Button(btn_frame, text="< Previous", command=self.prev_field).pack(side="left", padx=6)
        ttk.Button(btn_frame, text="Next >", command=self.next_field).pack(side="left")

        self.status_label = ttk.Label(frame, text="", foreground="green")
        self.status_label.pack(anchor="w", pady=(10, 0))

        ttk.Separator(frame).pack(fill="x", pady=14)

        auto_frame = ttk.Frame(frame)
        auto_frame.pack(fill="x")
        self.auto_btn = ttk.Button(auto_frame, text="⚡ Auto-fill ENTIRE agent",
                                    command=self.start_autofill)
        self.auto_btn.pack(side="left")

        self.auto_status_label = ttk.Label(frame, text="", foreground="#a45c00")
        self.auto_status_label.pack(anchor="w", pady=(6, 0))

        hint = ("Manual mode: Copy -> click the iWatch field -> Ctrl+V -> Next.\n"
                "Auto mode: click 'Auto-fill', then switch to iWatch and click on\n"
                "the 'First' field before the countdown ends. The program will\n"
                "paste and press Tab for you. It only pauses on Country/State so\n"
                "you can pick them by hand. Move the mouse to the top-left corner to abort.")
        ttk.Label(frame, text=hint, foreground="#555", justify="left").pack(anchor="w", pady=(14, 0))

    # ---------------- Logica ----------------
    def load_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm")])
        if not path:
            return
        try:
            self.wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            messagebox.showerror("Error", f"Could not open the file:\n{e}")
            return

        self.sheet_combo["values"] = self.wb.sheetnames
        if self.wb.sheetnames:
            self.sheet_combo.current(0)
            self.on_sheet_selected()

    def on_sheet_selected(self, event=None):
        if not self.wb:
            return
        sheet_name = self.sheet_var.get()
        ws = self.wb[sheet_name]
        self.blocks = extract_blocks_from_sheet(ws)

        self.role_combo["values"] = [b["name"] for b in self.blocks]
        if self.blocks:
            self.role_combo.current(0)
            self.on_role_selected()

    def on_role_selected(self, event=None):
        if not self.blocks:
            return
        idx = self.role_combo.current()
        if idx < 0:
            idx = 0
        self.fields = self.blocks[idx]["fields"]
        self.current_index = 0
        self.show_current_field()

    def show_current_field(self):
        if not self.fields:
            return
        label, value = self.fields[self.current_index]
        self.progress_label.config(
            text=f"Field {self.current_index + 1} of {len(self.fields)}  |  "
                 f"Agent: {self.sheet_var.get()}  |  Role: {self.role_var.get()}"
        )
        self.field_label.config(text=label)
        self.value_box.delete("1.0", "end")
        self.value_box.insert("1.0", value)
        self.status_label.config(text="" if value else "⚠ No value found for this field in the Excel")

    def copy_and_next(self):
        current_value = self.value_box.get("1.0", "end").strip()
        pyperclip.copy(current_value)
        self.status_label.config(text=f"✔ Copied: {current_value[:40]}")
        self.next_field()

    def next_field(self):
        if self.current_index < len(self.fields) - 1:
            self.current_index += 1
            self.show_current_field()
        else:
            self.status_label.config(text="✔ This was the last field for this agent")

    def prev_field(self):
        if self.current_index > 0:
            self.current_index -= 1
            self.show_current_field()

    # ---------------- Modo automatico ----------------
    def start_autofill(self):
        if not self.fields:
            messagebox.showwarning("No data", "First open an Excel file and choose an agent.")
            return

        confirm = messagebox.askyesno(
            "Confirm auto-fill",
            f"{len(self.fields)} fields will be filled automatically for:\n"
            f"Agent: '{self.sheet_var.get()}'\n"
            f"Role: '{self.role_var.get()}'\n\n"
            "Before continuing:\n"
            "1. Have iWatch open and visible, on a NEW 'Add Agent Contact' window.\n"
            "2. After clicking 'Yes', switch to iWatch and click on the\n"
            "   first field (First) BEFORE the countdown ends.\n\n"
            "Continue?"
        )
        if not confirm:
            return

        self.auto_btn.state(["disabled"])
        thread = threading.Thread(target=self._run_autofill, daemon=True)
        thread.start()

    def _run_autofill(self):
        # Countdown so the user has time to switch windows
        for i in range(6, 0, -1):
            self.after(0, lambda i=i: self.auto_status_label.config(
                text=f"Switch to iWatch and CLICK the 'First' field (make it active)... {i}"))
            time.sleep(1)

        self.after(0, lambda: self.auto_status_label.config(text="Auto-filling..."))

        # We minimize ourselves to GUARANTEE that Windows focus stays on
        # iWatch and not on this window (prevents the program from pasting
        # information into itself).
        self.after(0, self.iconify)
        time.sleep(0.5)

        for label, value in self.fields:
            if label in DROPDOWN_FIELDS:
                # Pause and wait for user confirmation for dropdown fields.
                # We restore the window and bring it to the front so the
                # message is ALWAYS visible (otherwise the script would be
                # stuck waiting for an OK you never see).
                done = threading.Event()

                def ask():
                    self.deiconify()
                    self.lift()
                    self.attributes("-topmost", True)
                    self.after(50, lambda: self.attributes("-topmost", False))
                    messagebox.showinfo(
                        "Select manually",
                        f"Select in the iWatch dropdown menu:\n\n"
                        f"{label} = {value or '(no value in Excel)'}\n\n"
                        f"Once selected, click OK to continue."
                    )
                    self.iconify()
                    done.set()

                self.after(0, ask)
                done.wait()
                time.sleep(0.5)
                pyautogui.press("tab")
                continue

            if value:
                pyperclip.copy(value)
                pyautogui.hotkey("ctrl", "v")
            pyautogui.press("tab")

        self.after(0, self.deiconify)
        self.after(0, lambda: self.auto_status_label.config(
            text="✔ Auto-fill complete. Carefully review iWatch before saving."))
        self.after(0, lambda: self.auto_btn.state(["!disabled"]))


if __name__ == "__main__":
    app = IWatchHelper()
    app.mainloop()
