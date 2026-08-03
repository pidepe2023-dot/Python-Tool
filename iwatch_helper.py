"""
Asistente de llenado iWatch
=============================
Lee un archivo Excel con varios tabs (uno por "master agente"), te deja
elegir cuál usar, y te va mostrando cada campo con un boton "Copiar" para
que lo pegues (Ctrl+V) directamente en el formulario de iWatch.

No hace clics automaticos en iWatch: tu controlas el mouse/teclado en todo
momento, asi que es seguro de usar aunque la ventana se mueva o cambie de
tamano, y no requiere permisos especiales de automatizacion de UI.

INSTALACION (en la compu del trabajo):
    pip install openpyxl pyperclip

USO:
    python iwatch_helper.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import pyperclip

# ---------------------------------------------------------------------------
# CONFIGURACION: ajusta esta lista al orden real de campos / Tab de iWatch.
# El "label" es como quieres verlo en pantalla; "keywords" son las palabras
# que el script buscara en la columna de etiquetas del Excel (no distingue
# mayusculas/minusculas) para encontrar el valor correspondiente.
# ---------------------------------------------------------------------------
FIELD_MAP = [
    {"label": "First",              "keywords": ["FIRST"]},
    {"label": "Middle",              "keywords": ["MIDDLE"]},
    {"label": "Last",                "keywords": ["LAST"]},
    {"label": "Suffix",              "keywords": ["SUFFIX"]},
    {"label": "Agency",              "keywords": ["AGENCY"]},
    {"label": "Address 1",           "keywords": ["ADDRESS 1", "ADDRESS1"]},
    {"label": "Address 2",           "keywords": ["ADDRESS 2", "ADDRESS2"]},
    {"label": "Country",             "keywords": ["COUNTRY"]},
    {"label": "State",               "keywords": ["STATE"]},
    {"label": "City",                "keywords": ["CITY"]},
    {"label": "Zip",                 "keywords": ["ZIP"]},
    {"label": "Source",              "keywords": ["SOURCE"]},
    {"label": "International Code",  "keywords": ["INTERNATIONAL CODE"]},
    {"label": "Phone Number",        "keywords": ["PHONE NUMBER"]},
    {"label": "Extension",           "keywords": ["EXTENSION"]},
    {"label": "Additional Phone",    "keywords": ["ADDITIONAL PHONE"]},
    {"label": "Fax Number",          "keywords": ["FAX"]},
    {"label": "Email Address",       "keywords": ["EMAIL"]},
]


def extract_fields_from_sheet(ws):
    """
    Recorre TODAS las celdas de la hoja buscando etiquetas conocidas
    (columna izquierda) y toma el valor de la celda vecina a la derecha.
    Esto lo hace robusto aunque cada tab tenga las filas en posiciones
    distintas (como en tu ejemplo, donde el bloque OWNER empieza en la
    fila 18 pero podria variar de tab a tab).
    """
    found = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value).strip().upper()
            if not text:
                continue
            for field in FIELD_MAP:
                if field["label"] in found:
                    continue
                for kw in field["keywords"]:
                    if kw == text or text.startswith(kw):
                        # Toma la celda inmediatamente a la derecha
                        neighbor = ws.cell(row=cell.row, column=cell.column + 1)
                        value = neighbor.value
                        if value is not None and str(value).strip() != "":
                            found[field["label"]] = str(value).strip()
                        break
    # Devuelve en el orden definido por FIELD_MAP, con "" si no se encontro
    return [(f["label"], found.get(f["label"], "")) for f in FIELD_MAP]


class IWatchHelper(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Asistente de llenado iWatch")
        self.geometry("480x420")
        self.resizable(False, False)

        self.wb = None
        self.fields = []  # lista de (label, value)
        self.current_index = 0

        self._build_top_bar()
        self._build_field_view()

    # ---------------- UI ----------------
    def _build_top_bar(self):
        frame = ttk.Frame(self, padding=10)
        frame.pack(fill="x")

        ttk.Button(frame, text="Abrir Excel...", command=self.load_excel).pack(side="left")

        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(frame, textvariable=self.sheet_var, state="readonly", width=30)
        self.sheet_combo.pack(side="left", padx=8)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_selected)

    def _build_field_view(self):
        frame = ttk.Frame(self, padding=20)
        frame.pack(fill="both", expand=True)

        self.progress_label = ttk.Label(frame, text="", font=("Segoe UI", 9))
        self.progress_label.pack(anchor="w")

        self.field_label = ttk.Label(frame, text="Selecciona un archivo y un agente para empezar",
                                      font=("Segoe UI", 14, "bold"))
        self.field_label.pack(anchor="w", pady=(10, 4))

        self.value_box = tk.Text(frame, height=3, font=("Segoe UI", 12), wrap="word")
        self.value_box.pack(fill="x", pady=(0, 12))

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill="x")

        self.copy_btn = ttk.Button(btn_frame, text="Copiar  (y avanzar)", command=self.copy_and_next)
        self.copy_btn.pack(side="left")

        ttk.Button(btn_frame, text="< Anterior", command=self.prev_field).pack(side="left", padx=6)
        ttk.Button(btn_frame, text="Siguiente >", command=self.next_field).pack(side="left")

        self.status_label = ttk.Label(frame, text="", foreground="green")
        self.status_label.pack(anchor="w", pady=(10, 0))

        hint = ("Tip: deja esta ventana y la de iWatch abiertas una junto a la otra.\n"
                "1) Clic en 'Copiar' -> 2) clic en el campo de iWatch -> 3) Ctrl+V -> 4) repite.")
        ttk.Label(frame, text=hint, foreground="#555", justify="left").pack(anchor="w", pady=(20, 0))

    # ---------------- Logica ----------------
    def load_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm")])
        if not path:
            return
        try:
            self.wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{e}")
            return

        self.sheet_combo["values"] = self.wb.sheetnames
        if self.wb.sheetnames:
            self.sheet_combo.current(0)
            self.on_sheet_selected()

    def on_sheet_selected(self, event=None):
        if not self.wb:
            return
        sheet_name = self.sheet_var.get()
        ws = self.wb[sheet_name]
        self.fields = extract_fields_from_sheet(ws)
        self.current_index = 0
        self.show_current_field()

    def show_current_field(self):
        if not self.fields:
            return
        label, value = self.fields[self.current_index]
        self.progress_label.config(
            text=f"Campo {self.current_index + 1} de {len(self.fields)}  |  Agente: {self.sheet_var.get()}"
        )
        self.field_label.config(text=label)
        self.value_box.delete("1.0", "end")
        self.value_box.insert("1.0", value)
        self.status_label.config(text="" if value else "⚠ No se encontro valor para este campo en el Excel")

    def copy_and_next(self):
        current_value = self.value_box.get("1.0", "end").strip()
        pyperclip.copy(current_value)
        self.status_label.config(text=f"✔ Copiado: {current_value[:40]}")
        self.next_field()

    def next_field(self):
        if self.current_index < len(self.fields) - 1:
            self.current_index += 1
            self.show_current_field()
        else:
            self.status_label.config(text="✔ Este era el ultimo campo del agente")

    def prev_field(self):
        if self.current_index > 0:
            self.current_index -= 1
            self.show_current_field()


if __name__ == "__main__":
    app = IWatchHelper()
    app.mainloop()
