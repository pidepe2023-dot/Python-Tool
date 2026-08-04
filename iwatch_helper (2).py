"""
Asistente de llenado iWatch
=============================
Lee un archivo Excel con varios tabs (uno por "master agente"). Dentro de
cada tab detecta automaticamente los distintos bloques de contacto (por
ejemplo OWNER, COMPLIANCE OFFICER) y te deja elegir cual usar. Luego te
va mostrando cada campo con un boton "Copiar" para que lo pegues (Ctrl+V)
en el formulario de iWatch, o usa el modo automatico para llenar todo.

INSTALACION (en la compu del trabajo):
    pip install openpyxl pyperclip pyautogui

USO:
    python iwatch_helper.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import pyperclip
import threading
import time
import pyautogui

# Failsafe de pyautogui: si mueves el mouse a la esquina superior izquierda
# de la pantalla en cualquier momento, se ABORTA el auto-llenado de emergencia.
pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.15  # pequena pausa entre cada accion simulada

# ---------------------------------------------------------------------------
# CONFIGURACION: ajusta esta lista al orden real de campos / Tab de iWatch.
# El "label" es como quieres verlo en pantalla; "keywords" son las palabras
# que el script buscara en la columna de etiquetas del Excel (no distingue
# mayusculas/minusculas) para encontrar el valor correspondiente.
# ---------------------------------------------------------------------------
FIELD_MAP = [
    {"label": "First",              "keywords": ["FIRST"]},
    {"label": "Middle",              "keywords": ["MIDDLE"]},
    {"label": "Last",                "keywords": ["LAST"]},
    {"label": "Suffix",              "keywords": ["SUFFIX"]},
    {"label": "Agency",              "keywords": ["AGENCY"]},
    {"label": "Address 1",           "keywords": ["ADDRESS 1", "ADDRESS1"]},
    {"label": "Address 2",           "keywords": ["ADDRESS 2", "ADDRESS2"]},
    {"label": "Country",             "keywords": ["COUNTRY"]},
    {"label": "State",               "keywords": ["STATE"]},
    {"label": "City",                "keywords": ["CITY"]},
    {"label": "Zip",                 "keywords": ["ZIP"]},
    {"label": "Source",              "keywords": ["SOURCE"]},
    {"label": "International Code",  "keywords": ["INTERNATIONAL CODE"]},
    {"label": "Phone Number",        "keywords": ["PHONE NUMBER"]},
    {"label": "Extension",           "keywords": ["EXTENSION"]},
    {"label": "Additional Phone",    "keywords": ["ADDITIONAL PHONE"]},
    {"label": "Fax Number",          "keywords": ["FAX"]},
    {"label": "Email Address",       "keywords": ["EMAIL"]},
]

# Campos que en iWatch son MENU DESPLEGABLE (combobox) en vez de texto libre.
# El auto-llenado se detiene en estos para que tu los selecciones a mano,
# porque pegar texto en un combobox no siempre funciona igual que en un
# campo de texto normal.
DROPDOWN_FIELDS = {"Country", "State"}


def _extract_fields_from_rows(ws, row_start, row_end):
    """Igual que antes, pero limitado a un rango de filas (un bloque)."""
    found = {}
    for row in ws.iter_rows(min_row=row_start, max_row=row_end):
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value).strip().upper()
            if not text:
                continue
            for field in FIELD_MAP:
                if field["label"] in found:
                    continue
                for kw in field["keywords"]:
                    if kw == text or text.startswith(kw):
                        neighbor = ws.cell(row=cell.row, column=cell.column + 1)
                        value = neighbor.value
                        if value is not None and str(value).strip() != "":
                            found[field["label"]] = str(value).strip()
                        break
    return [(f["label"], found.get(f["label"], "")) for f in FIELD_MAP]


# Todas las palabras clave conocidas (para saber que NO es un titulo de bloque)
_ALL_KEYWORDS = {kw for f in FIELD_MAP for kw in f["keywords"]}


def extract_blocks_from_sheet(ws):
    """
    Un mismo tab puede tener varios bloques de contacto (OWNER,
    COMPLIANCE OFFICER, etc.), cada uno con su propio FIRST/LAST/AGENCY...
    Esta funcion:
      1. Encuentra cada fila donde aparece la etiqueta "FIRST" -> ahi
         empieza un bloque nuevo.
      2. Busca hacia arriba el titulo del bloque (una celda en mayusculas
         que NO sea una de las etiquetas de campo conocidas, ej. "OWNER").
      3. Extrae los campos de ese bloque hasta la fila anterior al
         siguiente "FIRST" (o hasta el final de la hoja).
    Devuelve una lista de dicts: {"name": "OWNER", "fields": [...]}.
    Si no encuentra ningun "FIRST", regresa un solo bloque generico con
    todo el tab (para no romper tabs con un formato distinto).
    """
    first_rows = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value).strip().upper()
            if text == "FIRST":
                first_rows.append(cell.row)
                break

    if not first_rows:
        # No se detectaron bloques: usa el comportamiento anterior (todo el tab)
        return [{"name": "General", "fields": _extract_fields_from_rows(ws, 1, ws.max_row)}]

    blocks = []
    for i, start_row in enumerate(first_rows):
        end_row = (first_rows[i + 1] - 1) if i + 1 < len(first_rows) else ws.max_row

        # Busca el titulo del bloque: sube hasta 6 filas buscando texto en
        # mayusculas que no sea una etiqueta de campo conocida.
        block_name = f"Bloque {i + 1}"
        search_from = first_rows[i - 1] if i > 0 else 1
        for r in range(start_row - 1, max(start_row - 7, search_from - 1), -1):
            if r < 1:
                break
            for cell in ws[r]:
                if cell.value is None:
                    continue
                text = str(cell.value).strip()
                if not text or text.upper() in _ALL_KEYWORDS:
                    continue
                if text.upper() == text and len(text) > 2:  # todo mayusculas
                    block_name = text.title()
                    break
            if block_name != f"Bloque {i + 1}":
                break

        blocks.append({
            "name": block_name,
            "fields": _extract_fields_from_rows(ws, start_row, end_row),
        })

    return blocks


class IWatchHelper(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Asistente de llenado iWatch")
        self.geometry("520x600")
        self.minsize(480, 560)
        self.resizable(True, True)

        self.wb = None
        self.blocks = []  # lista de {"name":.., "fields": [...]}
        self.fields = []  # lista de (label, value) del bloque actual
        self.current_index = 0

        self._build_top_bar()
        self._build_field_view()

    # ---------------- UI ----------------
    def _build_top_bar(self):
        frame = ttk.Frame(self, padding=10)
        frame.pack(fill="x")

        ttk.Button(frame, text="Abrir Excel...", command=self.load_excel).pack(side="left")

        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(frame, textvariable=self.sheet_var, state="readonly", width=22)
        self.sheet_combo.pack(side="left", padx=8)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_selected)

        ttk.Label(frame, text="Rol:").pack(side="left", padx=(6, 2))
        self.role_var = tk.StringVar()
        self.role_combo = ttk.Combobox(frame, textvariable=self.role_var, state="readonly", width=18)
        self.role_combo.pack(side="left")
        self.role_combo.bind("<<ComboboxSelected>>", self.on_role_selected)

    def _build_field_view(self):
        frame = ttk.Frame(self, padding=20)
        frame.pack(fill="both", expand=True)

        self.progress_label = ttk.Label(frame, text="", font=("Segoe UI", 9))
        self.progress_label.pack(anchor="w")

        self.field_label = ttk.Label(frame, text="Selecciona un archivo y un agente para empezar",
                                      font=("Segoe UI", 14, "bold"))
        self.field_label.pack(anchor="w", pady=(10, 4))

        self.value_box = tk.Text(frame, height=3, font=("Segoe UI", 12), wrap="word")
        self.value_box.pack(fill="x", pady=(0, 12))

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill="x")

        self.copy_btn = ttk.Button(btn_frame, text="Copiar  (y avanzar)", command=self.copy_and_next)
        self.copy_btn.pack(side="left")

        ttk.Button(btn_frame, text="< Anterior", command=self.prev_field).pack(side="left", padx=6)
        ttk.Button(btn_frame, text="Siguiente >", command=self.next_field).pack(side="left")

        self.status_label = ttk.Label(frame, text="", foreground="green")
        self.status_label.pack(anchor="w", pady=(10, 0))

        ttk.Separator(frame).pack(fill="x", pady=14)

        auto_frame = ttk.Frame(frame)
        auto_frame.pack(fill="x")
        self.auto_btn = ttk.Button(auto_frame, text="⚡ Auto-llenar TODO el agente",
                                    command=self.start_autofill)
        self.auto_btn.pack(side="left")

        self.auto_status_label = ttk.Label(frame, text="", foreground="#a45c00")
        self.auto_status_label.pack(anchor="w", pady=(6, 0))

        hint = ("Modo manual: Copiar -> clic en el campo de iWatch -> Ctrl+V -> Siguiente.\n"
                "Modo automatico: clic en 'Auto-llenar', luego cambia a iWatch y haz clic\n"
                "en el campo 'First' antes de que termine la cuenta regresiva. El programa\n"
                "va a pegar y presionar Tab por ti. Se pausa solo en Country/State para que\n"
                "los elijas a mano. Mueve el mouse a la esquina sup. izquierda para abortar.")
        ttk.Label(frame, text=hint, foreground="#555", justify="left").pack(anchor="w", pady=(14, 0))

    # ---------------- Logica ----------------
    def load_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm")])
        if not path:
            return
        try:
            self.wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{e}")
            return

        self.sheet_combo["values"] = self.wb.sheetnames
        if self.wb.sheetnames:
            self.sheet_combo.current(0)
            self.on_sheet_selected()

    def on_sheet_selected(self, event=None):
        if not self.wb:
            return
        sheet_name = self.sheet_var.get()
        ws = self.wb[sheet_name]
        self.blocks = extract_blocks_from_sheet(ws)

        self.role_combo["values"] = [b["name"] for b in self.blocks]
        if self.blocks:
            self.role_combo.current(0)
            self.on_role_selected()

    def on_role_selected(self, event=None):
        if not self.blocks:
            return
        idx = self.role_combo.current()
        if idx < 0:
            idx = 0
        self.fields = self.blocks[idx]["fields"]
        self.current_index = 0
        self.show_current_field()

    def show_current_field(self):
        if not self.fields:
            return
        label, value = self.fields[self.current_index]
        self.progress_label.config(
            text=f"Campo {self.current_index + 1} de {len(self.fields)}  |  "
                 f"Agente: {self.sheet_var.get()}  |  Rol: {self.role_var.get()}"
        )
        self.field_label.config(text=label)
        self.value_box.delete("1.0", "end")
        self.value_box.insert("1.0", value)
        self.status_label.config(text="" if value else "⚠ No se encontro valor para este campo en el Excel")

    def copy_and_next(self):
        current_value = self.value_box.get("1.0", "end").strip()
        pyperclip.copy(current_value)
        self.status_label.config(text=f"✔ Copiado: {current_value[:40]}")
        self.next_field()

    def next_field(self):
        if self.current_index < len(self.fields) - 1:
            self.current_index += 1
            self.show_current_field()
        else:
            self.status_label.config(text="✔ Este era el ultimo campo del agente")

    def prev_field(self):
        if self.current_index > 0:
            self.current_index -= 1
            self.show_current_field()

    # ---------------- Modo automatico ----------------
    def start_autofill(self):
        if not self.fields:
            messagebox.showwarning("Sin datos", "Primero abre un Excel y elige un agente.")
            return

        confirm = messagebox.askyesno(
            "Confirmar auto-llenado",
            f"Se van a llenar {len(self.fields)} campos automaticamente para:\n"
            f"Agente: '{self.sheet_var.get()}'\n"
            f"Rol: '{self.role_var.get()}'\n\n"
            "Antes de continuar:\n"
            "1. Ten iWatch abierto y visible, en una ventana NUEVA de 'Add Agent Contact'.\n"
            "2. Despues de dar clic en 'Si', cambia a iWatch y haz clic en el\n"
            "   primer campo (First) ANTES de que termine la cuenta regresiva.\n\n"
            "¿Continuar?"
        )
        if not confirm:
            return

        self.auto_btn.state(["disabled"])
        thread = threading.Thread(target=self._run_autofill, daemon=True)
        thread.start()

    def _run_autofill(self):
        # Cuenta regresiva para que el usuario cambie de ventana a tiempo
        for i in range(6, 0, -1):
            self.after(0, lambda i=i: self.auto_status_label.config(
                text=f"Cambia a iWatch y haz CLIC en el campo 'First' (que quede activo)... {i}"))
            time.sleep(1)

        self.after(0, lambda: self.auto_status_label.config(text="Auto-llenando..."))

        # Nos minimizamos para GARANTIZAR que el foco de Windows quede en
        # iWatch y no en esta ventana (evita que el programa se pegue
        # informacion a si mismo).
        self.after(0, self.iconify)
        time.sleep(0.5)

        for label, value in self.fields:
            if label in DROPDOWN_FIELDS:
                # Pausa y espera confirmacion del usuario para campos de menu desplegable.
                # Restauramos la ventana y la traemos al frente para que el
                # mensaje SIEMPRE sea visible (si no, el script se queda
                # esperando un OK que nunca ves).
                done = threading.Event()

                def ask():
                    self.deiconify()
                    self.lift()
                    self.attributes("-topmost", True)
                    self.after(50, lambda: self.attributes("-topmost", False))
                    messagebox.showinfo(
                        "Selecciona manualmente",
                        f"Selecciona en el menu desplegable de iWatch:\n\n"
                        f"{label} = {value or '(sin valor en el Excel)'}\n\n"
                        f"Cuando lo hayas elegido, da clic en OK para continuar."
                    )
                    self.iconify()
                    done.set()

                self.after(0, ask)
                done.wait()
                time.sleep(0.5)
                pyautogui.press("tab")
                continue

            if value:
                pyperclip.copy(value)
                pyautogui.hotkey("ctrl", "v")
            pyautogui.press("tab")

        self.after(0, self.deiconify)
        self.after(0, lambda: self.auto_status_label.config(
            text="✔ Auto-llenado terminado. Revisa iWatch con cuidado antes de guardar."))
        self.after(0, lambda: self.auto_btn.state(["!disabled"]))


if __name__ == "__main__":
    app = IWatchHelper()
    app.mainloop()
